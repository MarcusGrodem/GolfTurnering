from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_DIR = Path(__file__).resolve().parent

NAVY = "17324D"
BLUE = "DCE6F1"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
RED = "F4CCCC"
WHITE = "FFFFFF"
GRAY = "E7E6E6"
THIN_GRAY = Side(style="thin", color="B7B7B7")
TABLE_BORDER = Border(bottom=THIN_GRAY)


def title(ws, text, subtitle=None):
    ws.merge_cells("A1:F1")
    ws["A1"] = text
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells("A2:F2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666")
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 32


def header_row(ws, row, labels):
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row, col, label)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(wrap_text=True)


def currency_cells(ws, cells):
    for cell in cells:
        ws[cell].number_format = '#,##0 "NOK"'


def setup_inputs(ws, values):
    title(ws, "Event Budget Inputs", "Edit yellow cells. Blue and green cells contain formulas or calculated outputs.")
    rows = [
        ("Event name", values.get("event_name", ""), "Name used in the workbook"),
        ("Venue", values.get("venue", ""), "Course or venue"),
        ("Total players", values.get("players", ""), "Maximum planned field"),
        ("Company share", values.get("company_share", 0.25), "Share of places included in company packages"),
        ("Company players", "=ROUND(B7*B8,0)", "Calculated"),
        ("Paying students", "=B7-B9", "Calculated"),
        ("Contingency rate", values.get("contingency", 0.10), "Applied to cash expenses"),
        ("Target cash sponsorship", values.get("sponsorship", ""), "Do not include in-kind prizes"),
        ("Provisional student fee", values.get("student_fee", ""), "Public registration price"),
    ]
    header_row(ws, 4, ["Input", "Value", "Notes"])
    for row, (label, value, note) in enumerate(rows, 5):
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.cell(row, 3, note)
        for col in range(1, 4):
            ws.cell(row, col).border = TABLE_BORDER
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=BLUE if isinstance(value, str) and value.startswith("=") else YELLOW)

    ws["B8"].number_format = "0%"
    ws["B11"].number_format = "0%"
    currency_cells(ws, ["B12", "B13"])

    header_row(ws, 16, ["Budget result", "Value", "Meaning"])
    outputs = [
        ("Expense subtotal", "=Costs!E19", "Before contingency"),
        ("Contingency", "=Costs!E20", "Reserve for uncertainty"),
        ("Working total", "=Costs!E21", "Total cash requirement"),
        ("Package revenue", "=Sponsorship!F10", "Revenue from package allocation"),
        ("Company places allocated", "=Sponsorship!E10", "Must not exceed company-player capacity"),
        ("Funding after package revenue", "=B19-B20", "Amount remaining for student fees or other income"),
        ("Break-even student fee", "=IF(B10=0,0,MAX(0,B22/B10))", "Required fee using package revenue"),
        ("Projected student revenue", "=B10*B13", "At full attendance"),
        ("Projected result", "=B20+B24-B19", "Positive is surplus; negative is funding gap"),
    ]
    for row, (label, formula, note) in enumerate(outputs, 17):
        ws.cell(row, 1, label)
        ws.cell(row, 2, formula)
        ws.cell(row, 3, note)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=GREEN)
        for col in range(1, 4):
            ws.cell(row, col).border = TABLE_BORDER
    currency_cells(ws, ["B17", "B18", "B19", "B20", "B22", "B23", "B24", "B25"])
    ws.conditional_formatting.add("B25", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add("B25", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor=GREEN)))
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 58


def setup_costs(ws, costs):
    title(ws, "Event Costs", "Enter quantities and unit costs. Use zero for sponsor-supplied items with no event cash expense.")
    header_row(ws, 4, ["Category", "Cost item", "Quantity", "Unit cost", "Total", "Status / notes"])
    for row, item in enumerate(costs, 5):
        category, name, quantity, unit_cost, note = item
        ws.cell(row, 1, category)
        ws.cell(row, 2, name)
        ws.cell(row, 3, quantity)
        ws.cell(row, 4, unit_cost)
        ws.cell(row, 5, f"=C{row}*D{row}")
        ws.cell(row, 6, note)
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(row, 4).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(row, 5).fill = PatternFill("solid", fgColor=BLUE)
        for col in range(1, 7):
            ws.cell(row, col).border = TABLE_BORDER

    header_row(ws, 19, ["", "Subtotal", "", "", "=SUM(E5:E18)", ""])
    ws["B20"] = "Contingency"
    ws["E20"] = "=E19*Inputs!B11"
    ws["B21"] = "Working total"
    ws["E21"] = "=E19+E20"
    for row in (20, 21):
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 5).font = Font(bold=True)
        ws.cell(row, 5).fill = PatternFill("solid", fgColor=GREEN)
    for row in range(5, 22):
        currency_cells(ws, [f"D{row}", f"E{row}"])
    ws.freeze_panes = "A5"
    widths = [18, 38, 12, 17, 19, 52]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def setup_sponsorship(ws, packages):
    title(ws, "Sponsorship Packages", "Package places are included benefits, not free event capacity. In-kind prizes are excluded from cash revenue.")
    header_row(ws, 4, ["Package", "Price", "Players included", "Packages sold", "Places allocated", "Cash revenue"])
    for row, item in enumerate(packages, 5):
        name, price, players, sold = item
        ws.cell(row, 1, name)
        ws.cell(row, 2, price)
        ws.cell(row, 3, players)
        ws.cell(row, 4, sold)
        ws.cell(row, 5, f"=C{row}*D{row}")
        ws.cell(row, 6, f"=B{row}*D{row}")
        for col in (2, 3, 4):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=YELLOW)
        for col in (5, 6):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=BLUE)
        for col in range(1, 7):
            ws.cell(row, col).border = TABLE_BORDER
    ws["A10"] = "TOTAL"
    ws["E10"] = "=SUM(E5:E9)"
    ws["F10"] = "=SUM(F5:F9)"
    for cell in ws[10]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=GREEN)
    ws["A12"] = "Company-player capacity"
    ws["E12"] = "=Inputs!B8"
    ws["A13"] = "Unallocated / overallocated places"
    ws["E13"] = "=E12-E10"
    ws["A14"] = "Direct event cost per player"
    ws["E14"] = "=SUMIF(Costs!B5:B18,\"Green fees\",Costs!D5:D18)+SUMIF(Costs!B5:B18,\"Dinner/food\",Costs!D5:D18)+SUMIF(Costs!B5:B18,\"On-course refreshments\",Costs!D5:D18)"
    currency_cells(ws, [f"B{row}" for row in range(5, 10)] + [f"F{row}" for row in range(5, 11)] + ["E14"])
    ws.conditional_formatting.add("E13", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)))
    ws.freeze_panes = "A5"
    for col, width in enumerate([28, 18, 19, 17, 20, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def setup_scenarios(ws):
    title(ws, "Student Fee Scenarios", "Change sponsorship values in column A to compare the required fee for paying students.")
    header_row(ws, 4, ["Cash sponsorship", "Working total", "Remaining funding", "Paying students", "Break-even fee per student", "Notes"])
    values = [0, 50000, 75000, 100000, 125000, 135000]
    for row, value in enumerate(values, 5):
        ws.cell(row, 1, value)
        ws.cell(row, 2, "=Costs!E21")
        ws.cell(row, 3, f"=MAX(0,B{row}-A{row})")
        ws.cell(row, 4, "=Inputs!B10")
        ws.cell(row, 5, f"=IF(D{row}=0,0,C{row}/D{row})")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=YELLOW)
        for col in (2, 3, 5):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=BLUE)
        for col in range(1, 7):
            ws.cell(row, col).border = TABLE_BORDER
        currency_cells(ws, [f"A{row}", f"B{row}", f"C{row}", f"E{row}"])
    ws.freeze_panes = "A5"
    for col, width in enumerate([22, 20, 22, 18, 28, 35], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build(path, values, costs, packages):
    wb = Workbook()
    wb.remove(wb.active)
    setup_inputs(wb.create_sheet("Inputs"), values)
    setup_costs(wb.create_sheet("Costs"), costs)
    setup_sponsorship(wb.create_sheet("Sponsorship"), packages)
    setup_scenarios(wb.create_sheet("Scenarios"))
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)


cost_names = [
    ("Venue", "Fixed venue charge"),
    ("Golf", "Green fees"),
    ("Food", "Dinner/food"),
    ("Food", "On-course refreshments"),
    ("Evening", "Dinner/afterparty room and AV"),
    ("Prizes", "Competition and main prizes"),
    ("Production", "Signs, badges, scorecards and stands"),
    ("Media", "Photographer/content"),
    ("Compliance", "Insurance, first aid and permits"),
    ("Digital", "Website, payment and administration"),
    ("Operations", "Event staff/volunteer expenses"),
    ("Logistics", "Local transport and miscellaneous"),
    ("Other", "Other expense 1"),
    ("Other", "Other expense 2"),
]

template_costs = [(category, name, "", "", "Enter estimate or confirmed quote") for category, name in cost_names]
template_packages = [("Package 1", "", "", ""), ("Package 2", "", "", ""), ("Package 3", "", "", ""), ("Package 4", "", "", ""), ("Package 5", "", "", "")]

sola_amounts = [
    (1, 17000, "Provided; confirm scope, VAT and terms"),
    (72, 900, "Provided; confirm VAT and minimum commitment"),
    (72, 450, "Working midpoint of NOK 400-500"),
    (72, 100, "Planning allowance"),
    (1, 10000, "Estimate; may be included in venue or catering"),
    (1, 0, "Non-cash items supplied by sponsors"),
    (1, 5000, "Planning allowance"),
    (1, 8000, "Optional planning allowance"),
    (1, 5000, "Confirm organizer and venue coverage"),
    (1, 5000, "Excludes development labor"),
    (1, 6000, "Confirm staffing plan"),
    (1, 3000, "Planning allowance"),
    (1, 0, "Unused allowance"),
    (1, 0, "Unused allowance"),
]
sola_costs = [
    (category, name, quantity, unit_cost, note)
    for (category, name), (quantity, unit_cost, note) in zip(cost_names, sola_amounts)
]
sola_packages = [
    ("Title Partner", 30000, 4, 1),
    ("Company Partner", 15000, 2, 5),
    ("Supporting Partner", 7500, 1, 4),
    ("Additional Package", 0, 0, 0),
    ("Additional Package", 0, 0, 0),
]

build(OUTPUT_DIR / "event-budget-template.xlsx", {}, template_costs, template_packages)
build(
    OUTPUT_DIR / "sola-gk-budget.xlsx",
    {
        "event_name": "Stavanger Student Golf Invitational",
        "venue": "Sola GK",
        "players": 72,
        "company_share": 0.25,
        "contingency": 0.10,
        "sponsorship": 135000,
        "student_fee": 900,
    },
    sola_costs,
    sola_packages,
)

for workbook_path in (OUTPUT_DIR / "event-budget-template.xlsx", OUTPUT_DIR / "sola-gk-budget.xlsx"):
    load_workbook(workbook_path)
    print(f"Created and validated {workbook_path.name}")
